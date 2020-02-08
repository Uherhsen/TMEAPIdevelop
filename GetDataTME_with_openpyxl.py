# -*- coding: utf-8 -*-
"""
Created on Tue Jan 14 12:02:49 2020

Получение данных о электронных компонентах с сайта TME через его API
"""
import openpyxl,time
from TME_Python_API import product_import_tme

# Функция считает все ячейки первого столбца в которых что то написано,до тех пор пока не встретит пустую ячейку "None"
def number_of_articles(path):
    # Открываем Эксель
    wb = openpyxl.load_workbook(path)#путь к файлу
    sheet = wb.active
    #выясняем количество артикулов в файле эксель
    i = 1
    while sheet['A'+str(i)].value != None:
        i+=1
    wb.save(path)
    return i-1

# Функция создающая список артикулов. Принимает число артикулов и номер колонки в виде буквы (str): 'A'- первая колонка
    
def articles_list(n,column,path):
    # Открываем Эксель
    wb = openpyxl.load_workbook(path)#путь к файлу
    sheet = wb.active 
    cord_in = column+str(1)
    cord_out = column+str((n)) 
    # формирование списка артикулов
    vals = [v[0].value for v in sheet[cord_in : cord_out]] #[r[0].value for r in sheet.Range(cord)]
    return vals
    wb.save(path)
    
# Функция для вывода пинга
    
def ping():
        ping_data = product_import_tme(token, app_secret, 'Utils/Ping', params={})
        print(ping_data)
        
# Функция проставляет оригинальные артикулы производлтеля, дескрипшен, ссылку на фото, ссылку на страницу продукта и вес в КГ! 
            
def search_articles(articles_list, path, rng1=0):
    '''Поиск базовых параметров'''
    # Открываем Эксель
    wb = openpyxl.load_workbook(path)#путь к файлу
    sheet = wb.active
    rng2=len(articles_list)
    for j in range(rng1,rng2):
        params['SearchPlain'] = str(articles_list[j])
        all_data = product_import_tme(token, app_secret, action1, params)
        #print(all_data)
        try :
            print(all_data['Data']['ProductList'][0]['Symbol'],all_data['Data']['ProductList'][0]['Description'])
            all_data['Data']['ProductList'][0]['Symbol']
            sheet['B'+str(j+1)] = all_data['Data']['ProductList'][0]['Symbol']
            sheet['C'+str(j+1)] = all_data['Data']['ProductList'][0]['Description']
            sheet['E'+str(j+1)] = all_data['Data']['ProductList'][0]['Photo'][2:]
            sheet['F'+str(j+1)] = all_data['Data']['ProductList'][0]['ProductInformationPage'][2:]
            weight = (all_data['Data']['ProductList'][0]['Weight'])
            if all_data['Data']['ProductList'][0]['WeightUnit']=='g':
                weight = weight*0.001
                #print('{:f}'.format(weight))
                weight = round(weight,(('{:f}'.format(weight)).count('0'))+1) # округление
            sheet['D'+str(j+1)] = weight
        except IndexError:
            if all_data["Status"]=="OK":
                print('\nСтатус сети ',all_data["Status"],'\nАртикул "'+articles_list[j]+'" отсутствует на TME\n')
                sheet['C'+str(j+1)] = 'Артикула нет на TME'
            else:
                print('\nСтатус сети ',all_data["Status"])
                sheet['C'+str(j+1)] = all_data['Data']['ProductList'][0]['Description']
                time.sleep(2) 
            continue
    wb.save(path)
    time.sleep(1)
    print('\nГотово')
    
# функция использует "экшен" для поиска параметров, к которому нужны оригинальные артикулы,
# запускается только после функции search_articles
    
def search_param(articles_list,path,rng1=0,):
    # Открываем Эксель
    wb = openpyxl.load_workbook(path)#путь к файлу
    sheet = wb.active
    rng2=len(articles_list)
    print('\nЦикл проставления параметров\n')
    for j in range(rng1,rng2):
        if articles_list[j] != None:
            params['SymbolList[0]'] = articles_list[j]
            all_data = product_import_tme(token, app_secret, action2, params)
            if all_data['Status'] == "OK":
                print(all_data['Data']['ProductList'][0]['ParameterList'][1]['ParameterName'],
                      all_data['Data']['ProductList'][0]['ParameterList'][1]['ParameterValue'])
                prms={}
                for i in all_data['Data']['ProductList'][0]['ParameterList']:
                    prms[i['ParameterName']] = i['ParameterValue']
                prms.pop('Производитель','Исключение')
                prms.pop('#Promotion','Исключение')
                sheet['G'+str(j+1)] = str(prms)
                
                time.sleep(0.3)
            else:
                print('ошибка статуса сети')
                j+=1
        else:
            print('Пропуск артикула')
            j+=1
    wb.save(path)
    time.sleep(1)
    print('\nГотово')

def products_files(articles_list, path, rng1=0):
    # Открываем Эксель
    wb = openpyxl.load_workbook(path)#путь к файлу
    sheet = wb.active
    rng2=len(articles_list)
    print('\nЦикл проставления ссылок на даташит\n')
    for j in range(rng1,rng2):
        if articles_list[j] != None:
            params['SymbolList[0]'] = articles_list[j]
            all_data = product_import_tme(token, app_secret, action3, params)
            if all_data['Status'] == "OK":
                try:
                    print(all_data['Data']['ProductList'][0]['Files']['DocumentList'][0]['DocumentUrl'][2:],'\n'+('_'*50)) #[0]['DocumentUrl'])
                    sheet['H'+str(j+1)] = str(all_data['Data']['ProductList'][0]['Files']['DocumentList'][0]['DocumentUrl'][2:])
                except KeyError:
                    print('KeyError')
                    j+=1
                except IndexError:
                    print('Нет PDF:\n',all_data['Data']['ProductList'][0]['Files'],'\n'+('_'*50))
                    j+=1
            else:
                print('ошибка статуса сети')
                break
                
        else:
            print('Нет артикула','\n'+('_'*50))
            j+=1
    print('\nГотово')
    wb.save(path)
    
    
if __name__ == '__main__': 
    
    xlsxpath = "D:\\python_programming\\TMEAPIdevelop_v7\\APP\\productdata.xlsx"
    
    params={'Country' : 'RU','Language' : 'RU',}
    token = 'ac434c181917ed4e51c49a2027bfd040e9f2da0054be7'
    app_secret = '0b748f6e5d340d693703'
    action1 = 'Products/Search' # request method, метод пинг Utils/Ping
    action2 = 'Products/GetParameters'
    action3 = 'Products/GetProductsFiles'
    
    n = number_of_articles(xlsxpath)
    
    work_articles_list_A = articles_list(n, 'A', xlsxpath)    
    search_articles(work_articles_list_A, xlsxpath)
    work_articles_list_B = articles_list(n,'B', xlsxpath)
    search_param(work_articles_list_B, xlsxpath)
    products_files(work_articles_list_B, xlsxpath)